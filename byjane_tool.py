import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
import io
import re

# --- 核心邏輯函數 ---
def f_prod_pack_col(value):
    mapping = {
        "ByJane 馬年限定禮盒 2026（宅配）": 1, 
        "ByJane 馬年限定禮盒 2026 （自取）": 1
    }
    return mapping.get(str(value).strip(), 0) if value else 0

def f_prod_type_col(value):
    mapping = {
        "Ａ ｜ The Medley Box  綜合風味": 3, "Ｂ｜ Waffle Lovers  人氣精選": 4,
        "Ｃ｜ The Refined Collection 成熟風味": 5, "Ｄ": 6,
        "經典原味": 7, "肉桂": 8, "濃心可可": 9, "藍莓乳酪": 10,
        "糖漬檸檬乳酪": 11, "芝麻": 12, "伯爵茶麻糬": 13, "抹茶紅豆麻糬": 14,
        "焙茶": 15, "培根楓糖起司": 16, "烤地瓜": 17, "焦糖杏仁奶油": 18,
        "鹽之花開心果可可": 19,"烤蒜檸香奶油": 20
    }
    return mapping.get(str(value).strip(), 0) if value else 0

# --- Streamlit 介面 ---
st.set_page_config(page_title="ByJane 訂單自動處理", layout="wide")
st.title("🧇 ByJane 訂單自動處理系統")

uploaded_file = st.file_uploader("請上傳『訂單報表』Excel 檔", type=["xlsx"])

if uploaded_file:
    # 提取日期
    file_name = uploaded_file.name
    date_suffix = ""
    date_match = re.search(r'(\d+)', file_name)
    if date_match:
        date_suffix = f"_{date_match.group(1)}"
    
    wb_source = load_workbook(uploaded_file, data_only=True)
    ws = wb_source.active
    
    header_map = {}
    for col in range(1, ws.max_column + 1):
        title = ws.cell(row=1, column=col).value
        if title: header_map[str(title).strip()] = col

    # --- 關鍵對齊：根據截圖修改標題搜尋關鍵字 ---
    col_id = header_map.get('訂單編號', 1)
    col_rcv_name = header_map.get('收件人名稱', 2)   # 修改為「名稱」
    col_rcv_mobile = header_map.get('收件人電話', 5) # 修改為「電話」
    col_rcv_addr = header_map.get('收件人地址', 7)   # 保持「地址」
    
    col_pay_stat = header_map.get('付款狀態', 7)
    col_pack = header_map.get('商品名稱', 9) 
    col_type = header_map.get('商品款式', 10) 
    col_val = header_map.get('數量', 11)
    col_deliver = header_map.get('配送方式', 14)

    pink_fill = PatternFill(start_color="FFC9CA", end_color="FFC9CA", fill_type="solid")

    # --- 數據收集 ---
    all_data = []
    active_cols = set()
    unpaid_list = []
    all_titles = ["訂單編號", "姓名", "A", "B", "C", "D", "原味", "肉桂", "可可", "藍莓", "檸檬", "芝麻", "伯爵", "抹茶", "焙茶", "培根", "地瓜", "焦糖", "開心果", "蒜檸"]
    
    row_source = 2
    order_num_flag = 0
    current_order_data = {}
    
    while True:
        order_num_current = ws.cell(row=row_source, column=col_id).value
        order_num_next = ws.cell(row=row_source + 1, column=col_id).value
        
        if order_num_flag == 0:
            p_status = str(ws.cell(row=row_source, column=col_pay_stat).value).strip()
            is_unpaid = (p_status == "等待付款")
            current_order_data = {
                "id": order_num_current,
                "name": ws.cell(row=row_source, column=col_rcv_name).value,
                "unpaid": is_unpaid,
                "items": {},
                "sum": 0,
                "deliver": ws.cell(row=row_source, column=col_deliver).value,
                "mobile": ws.cell(row=row_source, column=col_rcv_mobile).value,
                "address": ws.cell(row=row_source, column=col_rcv_addr).value
            }
            if is_unpaid:
                unpaid_list.append(f"🔴 {order_num_current} - {current_order_data['name']}")
            order_num_flag = 1

        p_val = ws.cell(row=row_source, column=col_val).value or 0
        if p_val > 0:
            p_col = f_prod_pack_col(ws.cell(row=row_source, column=col_pack).value)
            t_col = f_prod_type_col(ws.cell(row=row_source, column=col_type).value)
            target_col = p_col if p_col != 0 else t_col
            if 0 < target_col <= 20:
                current_order_data["items"][target_col] = current_order_data["items"].get(target_col, 0) + p_val
                active_cols.add(target_col)
                current_order_data["sum"] += (p_val * 8 if target_col < 7 else p_val)

        if order_num_current != order_num_next:
            all_data.append(current_order_data)
            order_num_flag = 0
        if order_num_next is None: break
        row_source += 1

    # --- 顯示未付款提醒 ---
    if unpaid_list:
        st.error("⚠️ 注意：以下訂單尚未付款，請核對後再出貨！")
        st.code("\n".join(unpaid_list))

    # --- 產出表格 ---
    used_indices = sorted(list(active_cols))
    final_header = ["訂單編號", "姓名"] + [all_titles[i-1] for i in used_indices if i > 2] + ["總數"]
    wb_byjane = Workbook(); ws_byjane = wb_byjane.active
    ws_byjane.append(final_header)

    wb_cat = Workbook(); ws_cat = wb_cat.active
    ws_cat.append(["收件人姓名", "收件人電話", "收件人手機", "收件人地址", "代收金額或到付", "件數", "品名(詳參數表)", "備註", "訂單編號", "希望配達時間(詳參數表)", "出貨日期(YYYY/MM/DD)", "預定配達日期(YYYY/MM/DD)", "溫層(詳參數表)", "尺寸(詳參數表)", "寄件人姓名", "寄件人電話", "寄件人手機", "寄件人地址", "保值金額", "品名說明", "是否列印(Y/N)", "是否捐贈(Y/N)", "統一編號", "手機載具", "愛心碼", "可刷卡(Y/N)", "手機支付(Y/N)"])

    wb_711 = Workbook(); ws_711 = wb_711.active
    ws_711.append(["訂單編號", "收件人姓名(必填)", "收件人手機(必填)", "FB名稱", "訂單備註", "代收金額", "門市編號(必填)", "匯款帳戶後五碼", "列印張數", "溫層(冷凍：0003)"])

    for data in all_data:
        # 宅配明細
        row_vals = [data["id"], data["name"]]
        for idx in used_indices:
            if idx > 2: row_vals.append(data["items"].get(idx, ""))
        row_vals.append(data["sum"] if data["sum"] > 0 else "")
        ws_byjane.append(row_vals)
        if data["unpaid"]:
            for c in range(1, len(row_vals) + 1):
                ws_byjane.cell(row=ws_byjane.max_row, column=c).fill = pink_fill

        # 填寫黑貓 & 711
        boxes = (data["sum"] // 90) + (1 if data["sum"] % 90 != 0 else 0) if data["sum"] > 0 else 1
        deliver = str(data["deliver"])
        
        if "黑貓冷凍宅配" in deliver:
            cat_row = [""] * 27
            cat_row[0], cat_row[1], cat_row[2], cat_row[3] = data["name"], data["mobile"], data["mobile"], data["address"]
            cat_row[5], cat_row[6], cat_row[8] = boxes, 1, data["id"]
            cat_row[9], cat_row[12], cat_row[13] = 4, 3, 1
            cat_row[14], cat_row[15], cat_row[16], cat_row[17] = "Byjane簡", "0960-319-998", "0960-319-998", "台南市中西區西賢五街26號"
            ws_cat.append(cat_row)
        elif "711" in deliver or "快速到店" in deliver:
            ws_711.append([data["id"], data["name"], data["mobile"], "", data["id"], "", "", "", boxes, "0003"])

    st.success(f"✨ 處理完成！")
    
    def get_io(wb):
        out = io.BytesIO(); wb.save(out); return out.getvalue()

    c1, c2, c3 = st.columns(3)
    c1.download_button(f"📂 宅配明細{date_suffix}", get_io(wb_byjane), f"宅配明細{date_suffix}.xlsx")
    c2.download_button(f"🚚 黑貓單{date_suffix}", get_io(wb_cat), f"黑貓單{date_suffix}.xlsx")
    c3.download_button(f"🏪 宅轉店{date_suffix}", get_io(wb_711), f"宅轉店{date_suffix}.xlsx")
